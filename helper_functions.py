from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import math
import shutil
import pandas as pd
import datetime
import numpy as np
from time import perf_counter
import json
from tabulate import tabulate

import load_params
import optim_model
import borefield_params
import eff_tables


def copy_file(src: str, dst: str) -> None:
    """
    Copies a file from src to dst.

    Args:
        src (str): Source file path.
        dst (str): Destination file path.

    Raises:
        FileNotFoundError: If the source file does not exist.
        PermissionError: If there are insufficient permissions to read or write.
        IOError: If there is an error during the copy operation.
    """
    try:
        shutil.copy(src, dst)
    except FileNotFoundError:
        raise FileNotFoundError(f"Source file not found: {src}")
    except PermissionError:
        raise PermissionError(
            f"Permission denied when copying file from {src} to {dst}"
        )
    except IOError as e:
        raise IOError(f"Error copying file from {src} to {dst}: {e}")


def change_model_name_in_mop(path_mop_file, old_name, new_name):
    """
    Updates the model name in a MOP file by replacing all occurrences of the old model name
    with a new model name in lines that start with 'optimization' or 'end'.

    Args:
        path_mop_file (str): The file path to the MOP file to be modified.
        old_name (str): The current model name to be replaced.
        new_name (str): The new model name to use as a replacement.

    Raises:
        FileNotFoundError: If the MOP file does not exist.
        PermissionError: If there are insufficient permissions to read or write the file.
        IOError: If there is an error reading or writing the file.
    """
    try:
        with open(path_mop_file, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        raise FileNotFoundError(f"MOP file not found: {path_mop_file}")
    except PermissionError:
        raise PermissionError(f"Permission denied when reading: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error reading file {path_mop_file}: {e}")

    for i, line in enumerate(lines):
        if f"optimization {old_name}" in line or f"end {old_name}" in line:
            lines[i] = line.replace(old_name, new_name)

    try:
        with open(path_mop_file, "w") as file:
            file.writelines(lines)
    except PermissionError:
        raise PermissionError(f"Permission denied when writing to: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error writing to file {path_mop_file}: {e}")


def load_json_file_as_dict(path):
    """
    Loads a JSON file and returns its content as a dictionary.

    Args:
        path (str): Path to the JSON file.

    Returns:
        dict: The content of the JSON file as a dictionary.

    Raises:
        FileNotFoundError: If the JSON file does not exist.
        PermissionError: If there are insufficient permissions to read the file.
        json.JSONDecodeError: If the file contains invalid JSON.
        IOError: If there is an error reading the file.
    """
    try:
        with open(path, "r", encoding="utf-8") as file:
            data = json.load(file)
    except FileNotFoundError:
        raise FileNotFoundError(f"JSON file not found: {path}")
    except PermissionError:
        raise PermissionError(f"Permission denied when reading: {path}")
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(
            f"Invalid JSON in file {path}: {e.msg}", e.doc, e.pos
        )
    except IOError as e:
        raise IOError(f"Error reading file {path}: {e}")

    return data


def change_weather_file_in_mop(path_mop_file, weather_file_name):
    """
    Updates the weather file name in a Modelica .mop file.

    Args:
        path_mop_file (str): The file path to the .mop file to be modified.
        weather_file_name (str): The name of the new weather file to reference in the .mop file.

    Raises:
        FileNotFoundError: If the MOP file does not exist.
        PermissionError: If there are insufficient permissions to read or write the file.
        IOError: If there is an error reading or writing the file.
        ValueError: If the weather file reference line is not found in the file.
    """
    try:
        with open(path_mop_file, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        raise FileNotFoundError(f"MOP file not found: {path_mop_file}")
    except PermissionError:
        raise PermissionError(f"Permission denied when reading: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error reading file {path_mop_file}: {e}")

    line_found = False
    for i, line in enumerate(lines):
        if "sim.filNam=" in line:
            lines[i] = (
                f'\t\t\t\t\t\t\t\t\tsim.filNam=Modelica.Utilities.Files.loadResource("modelica://IOCSmod/Resources/weatherdata//{weather_file_name}"),\n'
            )
            line_found = True
            break

    if not line_found:
        raise ValueError(
            f"Weather file reference line 'sim.filNam=' not found in {path_mop_file}"
        )

    try:
        with open(path_mop_file, "w") as file:
            file.writelines(lines)
    except PermissionError:
        raise PermissionError(f"Permission denied when writing to: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error writing to file {path_mop_file}: {e}")


def set_size_parameters_in_mop(path_mop_file, devices_info):
    """
    Updates size parameter values in a MOP file based on the provided devices_info dictionary.

    Args:
        path_mop_file (str): The file path to the MOP file to be updated.
        devices_info (dict): A nested dictionary containing device names as keys. Each device contains a "size_parameters" dictionary,
            where each key is a size variable and its value is a dictionary with:
                - "name_in_mop" (str): The variable name as it appears in the MOP file.
                - "value" (any): The new value to set for the variable.

    Raises:
        FileNotFoundError: If the MOP file does not exist.
        PermissionError: If there are insufficient permissions to read or write the file.
        IOError: If there is an error reading or writing the file.
        KeyError: If expected keys are missing in devices_info structure.
    """
    try:
        with open(path_mop_file, "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        raise FileNotFoundError(f"MOP file not found: {path_mop_file}")
    except PermissionError:
        raise PermissionError(f"Permission denied when reading: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error reading file {path_mop_file}: {e}")

    try:
        for i, line in enumerate(lines):
            for device, device_info in devices_info.items():
                for size_variable, variable_info in device_info[
                    "size_parameters"
                ].items():
                    if variable_info["name_in_mop"] in line:
                        lines[i] = (
                            f"\t\t\t\t\t\t\t\t\t{variable_info['name_in_mop']}{variable_info['value']},\n"
                        )
    except KeyError as e:
        raise KeyError(f"Missing expected key in devices_info structure: {e}")

    try:
        with open(path_mop_file, "w") as file:
            file.writelines(lines)
    except PermissionError:
        raise PermissionError(f"Permission denied when writing to: {path_mop_file}")
    except IOError as e:
        raise IOError(f"Error writing to file {path_mop_file}: {e}")


def update_computational_times(
    computational_times,
    write_inputs_time,
    load_params_time,
    size_optim_time,
    ocp_compilation_time,
    ocp_optimization_time,
    total_iteration_time,
):
    """
    Updates the computational times dictionary with new timing values.

    Args:
        computational_times (dict): Dictionary containing computational time entries with "value" keys.
        write_inputs_time (float): Time spent writing inputs.
        load_params_time (float): Time spent loading parameters.
        size_optim_time (float): Time spent on size optimization.
        ocp_compilation_time (float): Time spent on OCP compilation.
        ocp_optimization_time (float): Time spent on OCP optimization.
        total_iteration_time (float): Total iteration time.

    Returns:
        dict: Updated computational_times dictionary.

    Raises:
        KeyError: If expected keys are missing in computational_times dictionary.
        TypeError: If computational_times is not a dictionary.
    """
    if not isinstance(computational_times, dict):
        raise TypeError("computational_times must be a dictionary")

    try:
        computational_times["write_inputs_time"]["value"] = write_inputs_time
        computational_times["load_params_time"]["value"] = load_params_time
        computational_times["size_optim_time"]["value"] = size_optim_time
        computational_times["ocp_compilation_time"]["value"] = ocp_compilation_time
        computational_times["ocp_optimization_time"]["value"] = ocp_optimization_time
        computational_times["total_iteration_time"]["value"] = total_iteration_time
    except KeyError as e:
        raise KeyError(f"Missing expected key in computational_times: {e}")

    return computational_times


def read_ocp_result(path_outputsAll, path_OutputNames):
    """
    Reads OCP optimization results from CSV files and processes them into a DataFrame with time columns.

    Args:
        path_outputsAll (str): Path to the CSV file containing all outputs.
        path_OutputNames (str): Path to the file containing output column names.

    Returns:
        pd.DataFrame: DataFrame containing the results with columns including 'time', 'datetime', 'hours', and 'days'.

    Raises:
        FileNotFoundError: If either input file does not exist.
        pd.errors.ParserError: If there is an error parsing the CSV files.
        ValueError: If the data format is invalid.
    """
    try:
        column_names = pd.read_csv(path_OutputNames, header=None)
        column_names = column_names[0].tolist()
        column_names = ["time"] + column_names

        df = pd.read_csv(
            path_outputsAll,
            comment="#",
            skiprows=0,
            sep="\t",
            header=None,
            names=column_names,
        )
        df = df.iloc[1:]
        df = df.astype(float)

        start_date = datetime.datetime(2023, 1, 1)
        df["datetime"] = [
            start_date + datetime.timedelta(seconds=time) for time in df["time"]
        ]
        df["hours"] = [
            datetime.timedelta(seconds=time).total_seconds() / 3600
            for time in df["time"]
        ]
        df["days"] = [
            datetime.timedelta(seconds=time).total_seconds() / 86400
            for time in df["time"]
        ]
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Required file not found: {e}")
    except pd.errors.ParserError as e:
        raise pd.errors.ParserError(f"Error parsing CSV file: {e}")
    except (ValueError, KeyError) as e:
        raise ValueError(f"Invalid data format in result files: {e}")

    return df


def read_elec_use(df):
    """
    Calculates total electricity offtake and injection from a DataFrame containing electricity usage data.

    Args:
        df (pd.DataFrame): DataFrame containing an 'ElecUse' column with electricity usage values in W.

    Returns:
        tuple: (TotalOfftake, TotalInjection) in MWh, where:
            - TotalOfftake (float): Total electricity consumed from the grid.
            - TotalInjection (float): Total electricity injected into the grid.

    Raises:
        KeyError: If 'ElecUse' column is not found in the DataFrame.
        TypeError: If df is not a pandas DataFrame.
    """
    if not isinstance(df, pd.DataFrame):
        raise TypeError("Input must be a pandas DataFrame")

    try:
        elec_use = df["ElecUse"].to_numpy() / 1e3  # convert to kW
        elec_use_offtake = np.where(elec_use >= 0, elec_use, 0)
        elec_use_injection = np.where(elec_use < 0, -elec_use, 0)
        TotalOfftake = np.sum(elec_use_offtake) / 1e3  # convert to MWh
        TotalInjection = np.sum(elec_use_injection) / 1e3  # convert to MWh
    except KeyError:
        raise KeyError("'ElecUse' column not found in DataFrame")

    return TotalOfftake, TotalInjection


def read_discomfort(df):
    """
    Calculates total heating and cooling discomfort from a DataFrame.

    Args:
        df (pd.DataFrame): DataFrame containing 'DiscmfHea' and 'DiscmfCoo' columns with discomfort values.

    Returns:
        tuple: (TotHeaDiscmf, TotCooDiscmf) in Kh, where:
            - TotHeaDiscmf (float): Total heating discomfort.
            - TotCooDiscmf (float): Total cooling discomfort.

    Raises:
        KeyError: If required columns are not found in the DataFrame.
        TypeError: If df is not a pandas DataFrame.
    """
    if not isinstance(df, pd.DataFrame):
        raise TypeError("Input must be a pandas DataFrame")

    try:
        hea_discomfort = df["DiscmfHea"].to_numpy()
        coo_discomfort = df["DiscmfCoo"].to_numpy()
        TotHeaDiscmf = np.sum(hea_discomfort)  # Kh
        TotCooDiscmf = np.sum(coo_discomfort)  # Kh
    except KeyError as e:
        raise KeyError(f"Required column not found in DataFrame: {e}")

    return TotHeaDiscmf, TotCooDiscmf


def read_oper_variables_from_results(iteration_directory, operational_variables):
    """
    Reads operational variables from optimization results and updates the operational_variables dictionary.

    Args:
        iteration_directory (str or Path): Directory containing the result files.
        operational_variables (dict): Dictionary to store operational variables with structure:
            - "elec_offtake": {"value": float}
            - "elec_injection": {"value": float}
            - "hea_discmf": {"value": float}
            - "coo_discmf": {"value": float}

    Returns:
        dict: Updated operational_variables dictionary.

    Raises:
        FileNotFoundError: If result files are not found.
        KeyError: If expected keys are missing in operational_variables.
    """
    try:
        df = read_ocp_result(
            f"{iteration_directory}/outputsAll.csv",
            f"{iteration_directory}/OutputNames.txt",
        )
        (
            operational_variables["elec_offtake"]["value"],
            operational_variables["elec_injection"]["value"],
        ) = read_elec_use(df)
        (
            operational_variables["hea_discmf"]["value"],
            operational_variables["coo_discmf"]["value"],
        ) = read_discomfort(df)
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Result files not found in {iteration_directory}: {e}")
    except KeyError as e:
        raise KeyError(f"Missing expected key in operational_variables: {e}")

    return operational_variables


def update_overview_sheet(
    path_overview_sheet,
    devices_info,
    operational_variables,
    capex,
    opex,
    maintCost,
    computational_times,
    iteration,
):
    """
    Updates an Excel overview sheet with device sizing, operational variables, costs, and computational times.

    Args:
        path_overview_sheet (str): Path to the Excel file.
        devices_info (dict): Dictionary containing device information and size parameters.
        operational_variables (dict): Dictionary containing operational variables and their Excel row positions.
        capex (float): Capital expenditure value.
        opex (float): Operating expenditure value.
        maintCost (float): Maintenance cost value.
        computational_times (dict): Dictionary containing computational times and their Excel row positions.
        iteration (int): Current iteration number (determines the column offset).

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        PermissionError: If there are insufficient permissions to read or write the file.
        KeyError: If expected keys are missing in the dictionaries.
    """
    try:
        wb = load_workbook(path_overview_sheet)
        sheet = wb["Overview"]
    except FileNotFoundError:
        raise FileNotFoundError(f"Overview sheet not found: {path_overview_sheet}")
    except PermissionError:
        raise PermissionError(
            f"Permission denied when accessing: {path_overview_sheet}"
        )
    except KeyError:
        raise KeyError("'Overview' sheet not found in workbook")

    try:
        first_iteration_column = 3  # C
        col_letter = get_column_letter(first_iteration_column + iteration)

        for device, device_info in devices_info.items():
            for variable, variable_info in device_info["size_parameters"].items():
                if variable_info["excel_row"] == "X":
                    continue
                cell = col_letter + str(variable_info["excel_row"])
                sheet[cell] = variable_info["value"]

        for variable, value in operational_variables.items():
            cell = col_letter + str(operational_variables[variable]["excel_row"])
            sheet[cell] = operational_variables[variable]["value"]

        # Fill in capex, opex, maintCost, and TCO
        sheet[col_letter + "22"] = capex
        sheet[col_letter + "23"] = opex
        sheet[col_letter + "24"] = maintCost
        sheet[col_letter + "25"] = capex + opex + maintCost

        # Write computational times
        for time_key, time_info in computational_times.items():
            cell = col_letter + str(time_info["excel_row"])
            sheet[cell] = time_info["value"]

        wb.save(path_overview_sheet)
    except KeyError as e:
        raise KeyError(f"Missing expected key in input dictionaries: {e}")
    except Exception as e:
        raise Exception(f"Error updating overview sheet: {e}")


def convert_numpy_to_list_sizing_result_dict(obj):
    """
    Recursively converts all numpy arrays in a dictionary (or list) to lists.

    Args:
        obj (dict, list, or other): The object to process.

    Returns:
        The processed object with numpy arrays converted to lists.
    """
    if isinstance(obj, dict):
        return {
            key: convert_numpy_to_list_sizing_result_dict(value)
            for key, value in obj.items()
        }
    elif isinstance(obj, list):
        return [convert_numpy_to_list_sizing_result_dict(item) for item in obj]
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    else:
        return obj


def save_sizing_result_dict_to_json(result_dict, iteration_directory):
    """
    Saves a sizing result dictionary to a JSON file after converting numpy arrays to lists.

    Args:
        result_dict (dict): The sizing result dictionary to save.
        iteration_directory (str or Path): The directory where the JSON file will be saved.

    Raises:
        IOError: If there is an error writing to the file.
        TypeError: If result_dict is not serializable to JSON.
    """
    try:
        result_dict = convert_numpy_to_list_sizing_result_dict(result_dict)
        with open(iteration_directory / "sizing_result_dict.json", "w") as file:
            json.dump(result_dict, file, indent=4)
    except IOError as e:
        raise IOError(f"Error writing sizing result dict to file: {e}")
    except TypeError as e:
        raise TypeError(f"Result dict contains non-serializable data: {e}")


def save_devices_info_to_json(devices_info, iteration_directory):
    """
    Saves the devices_info dictionary to a JSON file.

    Args:
        devices_info (dict): The devices_info dictionary to save.
        iteration_directory (str or Path): The directory where the JSON file will be saved.

    Raises:
        IOError: If there is an error writing to the file.
        TypeError: If devices_info is not serializable to JSON.
    """
    try:
        with open(iteration_directory / "devices_info.json", "w") as file:
            json.dump(devices_info, file, indent=4)
    except IOError as e:
        raise IOError(f"Error writing devices info to file: {e}")
    except TypeError as e:
        raise TypeError(f"Devices info contains non-serializable data: {e}")


def run_size_optim():
    """
    Runs the size optimization process including parameter loading and iterative borefield sizing.

    Returns:
        tuple: (result_dict, load_params_time, size_optim_time) where:
            - result_dict (dict): Dictionary containing optimization results.
            - load_params_time (float): Time spent loading parameters in seconds.
            - size_optim_time (float): Time spent on size optimization in seconds.

    Raises:
        Exception: If there are errors during parameter loading or optimization.
    """
    try:
        # Load parameters
        start = perf_counter()
        param, devs, dem, result_dict = load_params.load_params()
        end = perf_counter()
        load_params_time = end - start

        # Run optimization
        start = perf_counter()

        Borefield_sizing_finished = False
        while not Borefield_sizing_finished:
            # Set max gshp cap based on borefield sizing
            result_dict = optim_model.run_optim(devs, param, dem, result_dict)
            if result_dict["Borefield"]["cap"] == 0:
                Borefield_sizing_finished = True
            elif (
                result_dict["Borefield"]["nBor"] == 2
                and not result_dict["Borefield"]["Sized_to_max"]
            ):
                Borefield_sizing_finished = True
            elif (
                result_dict["Borefield"]["Sized_to_min"] == False
                and result_dict["Borefield"]["depth"] > 50
                and not result_dict["Borefield"]["Sized_to_max"]
            ):
                Borefield_sizing_finished = True
            elif result_dict["Borefield"]["Sized_to_max"] == True:
                N_1 = devs["Borefield"]["bor_params"]["N_1"]
                N_2 = devs["Borefield"]["bor_params"]["N_2"]
                N_1, N_2 = borefield_params.increase_nBor(N_1, N_2)
                borefield_params.write_N_1_N_2_to_json(devs, N_1, N_2)
                devs = borefield_params.set_general_parameters(devs)
                devs = borefield_params.set_calculation_parameters(devs)
                print(
                    f"Sized to maximum allowable borefield size: increasing nBor to {N_1*N_2} ({N_1}x{N_2})"
                )
            else:
                N_1 = devs["Borefield"]["bor_params"]["N_1"]
                N_2 = devs["Borefield"]["bor_params"]["N_2"]
                N_1, N_2, nBor = borefield_params.find_next_N_1_N_2(N_1, N_2)
                borefield_params.write_N_1_N_2_to_json(devs, N_1, N_2)
                devs = borefield_params.set_general_parameters(devs)
                devs = borefield_params.set_calculation_parameters(devs)
                print(
                    f"Sized to minimum allowable borefield size: decrasing nBor to {N_1*N_2} ({N_1}x{N_2})"
                )
        end = perf_counter()

        size_optim_time = end - start
        print(f"Size optimization time: {size_optim_time} s")

        return result_dict, load_params_time, size_optim_time
    except Exception as e:
        raise Exception(f"Error during size optimization: {e}")


def print_sizing_result_overview(result_dict):
    """
    Prints a formatted overview of sizing optimization results including general KPIs and device-specific results.

    Args:
        result_dict (dict): Dictionary containing optimization results with keys for general KPIs,
            installed devices, and device-specific results.

    Raises:
        KeyError: If expected keys are missing in result_dict.
    """
    try:
        # General
        print("---------- General Results ----------")
        table = [
            ["tac", result_dict["tac"], "euro"],
            ["co2", result_dict["co2"], "tonCO2/year"],
            ["CAPEX initial", result_dict["CAPEX_initial"], "euro"],
            ["CAPEX annualized", result_dict["CAPEX_annual"], "euro/year"],
            ["OPEX yearly total", result_dict["OPEX_yearly"], "euro/year"],
            [
                "OPEX yearly energy and CO2",
                result_dict["OPEX_yearly_energy"],
                "euro/year",
            ],
            ["OPEX yearly maintenance", result_dict["OPEX_yearly_maint"], "euro/year"],
            ["Electricity offtake", result_dict["from_el_grid_total"], "MWh/year"],
            ["Electricity injection", result_dict["to_el_grid_total"], "MWh/year"],
        ]

        print(tabulate(table, headers=["KPI", "Value", "Unit"], tablefmt="grid"))

        # Devices
        for dev in result_dict["installed_devices"]:
            print("---------- Results", dev, "----------")
            for key in result_dict[dev]:
                print(key, ": ", result_dict[dev][key])
    except KeyError as e:
        raise KeyError(f"Missing expected key in result_dict: {e}")


def read_optimal_sizes(result_dict, devices_info):
    """
    Reads optimal sizes from optimization results and updates the devices_info dictionary.

    Args:
        result_dict (dict): Dictionary containing optimization results for each device.
        devices_info (dict): Dictionary containing device information to be updated with optimal sizes.

    Returns:
        dict: Updated devices_info dictionary with optimal size parameters.

    Raises:
        KeyError: If expected keys are missing in dictionaries.
    """
    try:
        for device, device_info in devices_info.items():
            if device in ["ASHP", "GSHP", "PV", "STC", "PVT", "ASCHI"]:
                if result_dict[device]["cap"] > 0.01:
                    devices_info[device]["size_parameters"]["Size"]["value"] = (
                        result_dict[device]["cap"]
                    )
                else:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0.01
            if device in ["Borefield"]:
                devices_info[device]["size_parameters"]["Size"]["value"] = result_dict[
                    device
                ]["cap"]
                devices_info[device]["size_parameters"]["TBorMin"]["value"] = (
                    result_dict[device]["Tf_min"] + 273.15
                )
                devices_info[device]["size_parameters"]["rCel"]["value"] = result_dict[
                    device
                ]["rCel"]
                devices_info[device]["size_parameters"]["kappa"]["value"] = result_dict[
                    device
                ]["kappa"]
                if result_dict[device]["cap"] == 0:
                    devices_info[device]["size_parameters"]["depth"]["value"] = 1
                    devices_info[device]["size_parameters"]["nBor"]["value"] = int(1)
                else:
                    devices_info[device]["size_parameters"]["depth"]["value"] = (
                        result_dict[device]["depth"]
                    )
                    devices_info[device]["size_parameters"]["nBor"]["value"] = int(
                        result_dict[device]["nBor"]
                    )

            if device in ["TES", "CTES"]:
                if result_dict[device]["volume"] > 0.01:
                    devices_info[device]["size_parameters"]["Size"]["value"] = (
                        result_dict[device]["volume"]
                    )
                else:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0.01
            if device in ["BAT"]:
                if result_dict[device]["cap"] > 0.01:
                    devices_info[device]["size_parameters"]["Size"]["value"] = (
                        result_dict[device]["cap"]
                    )
                    devices_info[device]["size_parameters"]["hasBat"]["value"] = "true"
                    devices_info[device]["size_parameters"]["PCha_max"]["value"] = (
                        result_dict[device]["cap"] / 2
                    )
                else:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0.01
                    devices_info[device]["size_parameters"]["hasBat"]["value"] = "false"
                    devices_info[device]["size_parameters"]["PCha_max"]["value"] = 0.01
    except KeyError as e:
        raise KeyError(f"Missing expected key in result_dict or devices_info: {e}")

    return devices_info


def set_sizes_of_non_feasible_devices_to_zero(devices_info):
    """
    Sets the sizes of non-feasible devices to zero in the devices_info dictionary.

    Args:
        devices_info (dict): Dictionary containing device size variables and feasibility flags.

    Returns:
        dict: Updated devices_info dictionary with non-feasible device sizes set to zero.

    Raises:
        KeyError: If expected keys are missing in devices_info structure.
    """
    try:
        for device, device_info in devices_info.items():
            if not device_info["feasible"]:
                if device in [
                    "ASHP",
                    "GSHP",
                    "PV",
                    "STC",
                    "PVT",
                    "TES",
                    "CTES",
                    "ASCHI",
                ]:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0.01
                if device in ["Borefield"]:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0
                    devices_info[device]["size_parameters"]["depth"]["value"] = 1
                    devices_info[device]["size_parameters"]["nBor"]["value"] = int(1)
                if device in ["BAT"]:
                    devices_info[device]["size_parameters"]["Size"]["value"] = 0.01
                    devices_info[device]["size_parameters"]["hasBat"]["value"] = "false"
                    devices_info[device]["size_parameters"]["PCha_max"]["value"] = 0.01
    except KeyError as e:
        raise KeyError(f"Missing expected key in devices_info: {e}")

    return devices_info


def write_input_profiles(path_results, path_input_data):
    """
    Generates and writes input profiles (demands, COPs, EERs) from optimization results.

    Args:
        path_results (Path): Path to the directory containing optimization results.
        path_input_data (Path): Path to the directory where input profiles and COP tables are stored.

    Returns:
        float: Time spent writing input profiles in seconds.

    Raises:
        FileNotFoundError: If required result files or COP tables are not found.
        IOError: If there is an error writing to output files.
    """
    try:
        start = perf_counter()

        # Read data
        path_outputsAll = path_results / "outputsAll.csv"
        path_OutputNames = path_results / "OutputNames.txt"
        df = read_ocp_result(path_outputsAll, path_OutputNames)

        # Load demand profiles
        QDem = (df["QDem"]) / 1e3
        QDemHea = np.where(QDem >= 0, QDem, 0)
        QDemCoo = np.where(QDem < -0.01, -QDem, 0)
        ElecDem = df["PDem"].to_numpy() / 1e3

        QBor = df["QBor"].to_numpy()
        inj = np.where(QBor > 0, QBor, 0) / 1e3
        ext = np.where(QBor < 0, -QBor, 0) / 1e3

        # Calculate COPs
        TSup = df["T_COP_Sup"].to_numpy()
        Tair = df["T_COP_Air"].to_numpy()
        TBor = df["T_COP_Bor"].to_numpy()

        ASHP_table_path = path_input_data / "COPTables" / "COPTable_VitoCal200A.csv"
        GSHP_table_path = (
            path_input_data / "COPTables" / "COPTable_VitoCal300G_BWS_A21.csv"
        )
        CHI_table_path = (
            path_input_data / "COPTables" / "EERTable_Galetti_MPE013H0AA.csv"
        )

        cop_gshp = eff_tables.eff(eff_table_path=GSHP_table_path)
        cop_ashp = eff_tables.eff(eff_table_path=ASHP_table_path)
        eer_chi = eff_tables.eff(eff_table_path=CHI_table_path)

        print("Calculating COPs and EERs...")
        COP_ASHP = np.zeros(len(TSup))
        COP_GSHP = np.zeros(len(TSup))
        EER_CHI = np.zeros(len(TSup))

        for i in range(len(TSup)):
            COP_ASHP[i] = cop_ashp.get_eff(Tsink=TSup[i], Tsource=Tair[i])
            COP_GSHP[i] = cop_gshp.get_eff(Tsink=TSup[i], Tsource=TBor[i])
            EER_CHI[i] = eer_chi.get_eff(Tsink=Tair[i], Tsource=TSup[i])
        print("COPs and EERs calculated!")

        path_input_profiles = path_input_data / "Profiles"
        np.savetxt(path_input_profiles / "heating_demand.txt", QDemHea, fmt="%.1f")
        np.savetxt(path_input_profiles / "cooling_demand.txt", QDemCoo, fmt="%.1f")
        np.savetxt(path_input_profiles / "electricity_demand.txt", ElecDem, fmt="%.1f")
        np.savetxt(path_input_profiles / "COP_ASHP.txt", COP_ASHP, fmt="%.1f")
        np.savetxt(path_input_profiles / "COP_GSHP.txt", COP_GSHP, fmt="%.1f")
        np.savetxt(path_input_profiles / "EER_ASCHI.txt", EER_CHI, fmt="%.1f")
        np.savetxt(path_input_profiles / "initial_borefield_ext.txt", ext, fmt="%.1f")
        np.savetxt(path_input_profiles / "initial_borefield_inj.txt", inj, fmt="%.1f")

        end = perf_counter()
        write_inputs_time = end - start
        print(f"Write input profiles time: {write_inputs_time} s")

        return write_inputs_time
    except FileNotFoundError as e:
        raise FileNotFoundError(f"Required file not found: {e}")
    except IOError as e:
        raise IOError(f"Error writing input profiles: {e}")


def calculate_NPV_factor_inv_costs(interest_rate, life_time, observation_time):
    """
    Calculates the Net Present Value (NPV) factor for investment costs considering replacements and residual value.

    Args:
        interest_rate (float): Interest rate (as a decimal, e.g., 0.05 for 5%).
        life_time (float): Lifetime of the investment in years.
        observation_time (float): Observation period in years.

    Returns:
        float: The NPV factor accounting for initial investment, replacements, and residual value.

    Raises:
        ValueError: If interest_rate, life_time, or observation_time are invalid.
    """
    if interest_rate < 0:
        raise ValueError("Interest rate must be non-negative")
    if life_time <= 0:
        raise ValueError("Life time must be positive")
    if observation_time <= 0:
        raise ValueError("Observation time must be positive")

    try:
        q = 1 + interest_rate
        n = int(math.floor(observation_time / life_time))
        invest_replacements = sum((q ** (-i * life_time)) for i in range(1, n + 1))
        res_value = (
            ((n + 1) * life_time - observation_time)
            / life_time
            * (q ** (-observation_time))
        )
        NPV_factor = 1 + invest_replacements - res_value
    except (ZeroDivisionError, OverflowError) as e:
        raise ValueError(f"Error calculating NPV factor: {e}")

    return NPV_factor


def add_NPV_of_rel_inv_cost_to_devices_info(
    devices_info, interest_rate, observation_time
):
    """
    Adds the Net Present Value (NPV) of relative investment costs to each device in devices_info.

    Args:
        devices_info (dict): Dictionary containing device information with 'life_time' and 'inv_cost' keys.
        interest_rate (float): Interest rate (as a decimal, e.g., 0.05 for 5%).
        observation_time (float): Observation period in years.

    Returns:
        dict: Updated devices_info dictionary with 'NPV_inv_cost' added for each device.

    Raises:
        KeyError: If required keys are missing in devices_info.
        ValueError: If interest_rate or observation_time are invalid.
    """
    try:
        for device, device_info in devices_info.items():
            NPV_factor = calculate_NPV_factor_inv_costs(
                interest_rate, device_info["life_time"], observation_time
            )
            devices_info[device]["NPV_inv_cost"] = NPV_factor * device_info["inv_cost"]
    except KeyError as e:
        raise KeyError(f"Missing expected key in devices_info: {e}")

    return devices_info


def calculate_capex(devices_info, interest_rate, observation_time):
    """
    Calculates the total capital expenditure (CAPEX) for all devices.

    Args:
        devices_info (dict): Dictionary containing device information with 'life_time', 'inv_cost', and
            'size_parameters' keys.
        interest_rate (float): Interest rate (as a decimal, e.g., 0.05 for 5%).
        observation_time (float): Observation period in years.

    Returns:
        tuple: (capex, devices_info) where:
            - capex (float): Total capital expenditure for all devices.
            - devices_info (dict): Updated dictionary with NPV costs and CAPEX for each device.

    Raises:
        KeyError: If expected keys are missing in devices_info.
    """
    try:
        capex = 0
        for device, device_info in devices_info.items():
            devices_info[device]["capex"] = (
                devices_info[device]["NPV_inv_cost"]
                * device_info["size_parameters"]["Size"]["value"]
            )
            capex += devices_info[device]["capex"]
    except KeyError as e:
        raise KeyError(f"Missing expected key in devices_info: {e}")

    return capex, devices_info


def calculate_opex_and_maintCost(
    operational_variables,
    devices_info,
    interest_rate,
    observation_time,
    elec_price_offtake,
    elec_price_injection,
):
    """
    Calculates the operational expenditure (OPEX) and maintenance costs.

    Args:
        operational_variables (dict): Dictionary containing electricity offtake and injection values.
        devices_info (dict): Dictionary containing device information including maintenance costs.
        interest_rate (float): Interest rate (as a decimal, e.g., 0.05 for 5%).
        observation_time (float): Observation period in years.
        elec_price_offtake (float): Electricity price for offtake in €/MWh.
        elec_price_injection (float): Electricity price for injection in €/MWh.

    Returns:
        tuple: (opex, maintCost, devices_info) where:
            - opex (float): Total operational expenditure.
            - maintCost (float): Total maintenance cost.
            - devices_info (dict): Updated dictionary with maintenance costs for each device.

    Raises:
        KeyError: If expected keys are missing in dictionaries.
        ValueError: If interest_rate or observation_time are invalid.
    """
    if interest_rate <= 0:
        raise ValueError("Interest rate must be positive")
    if observation_time <= 0:
        raise ValueError("Observation time must be positive")

    try:
        annuity_factor = (1 / interest_rate) * (
            1 - (1 / (1 + interest_rate) ** observation_time)
        )

        # Calculate OPEX
        opex = (
            operational_variables["elec_offtake"]["value"]
            * elec_price_offtake
            * annuity_factor
        )
        opex -= (
            operational_variables["elec_injection"]["value"]
            * elec_price_injection
            * annuity_factor
        )

        # Calculate maintCost
        maintCost = 0
        for device, device_info in devices_info.items():
            devices_info[device]["maintCost"] = (
                device_info["cost_om"]
                * device_info["inv_cost"]
                * device_info["size_parameters"]["Size"]["value"]
            )
            maintCost += devices_info[device]["maintCost"]

        maintCost *= annuity_factor
    except KeyError as e:
        raise KeyError(f"Missing expected key in input dictionaries: {e}")

    return opex, maintCost, devices_info


def calculate_maintCost(devices_info, interest_rate, observation_time):
    """
    Calculates the total maintenance costs over the observation period.

    Args:
        devices_info (dict): Dictionary containing device information including 'cost_om' and 'size_parameters'.
        interest_rate (float): Interest rate (as a decimal, e.g., 0.05 for 5%).
        observation_time (float): Observation period in years.

    Returns:
        float: Total maintenance costs over the observation period.

    Raises:
        KeyError: If expected keys are missing in devices_info.
        ValueError: If interest_rate or observation_time are invalid.
    """
    if interest_rate <= 0:
        raise ValueError("Interest rate must be positive")
    if observation_time <= 0:
        raise ValueError("Observation time must be positive")

    try:
        maintCost = 0
        for device, device_info in devices_info.items():
            devices_info[device]["maintCost"] = (
                device_info["cost_om"] * device_info["size_parameters"]["Size"]["value"]
            )
            maintCost += devices_info[device]["maintCost"]
    except KeyError as e:
        raise KeyError(f"Missing expected key in devices_info: {e}")

    return maintCost
